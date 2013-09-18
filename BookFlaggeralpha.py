from openpyxl import load_workbook
from openpyxl import Workbook
import os
import sys
import re
from book import LibraryBook
import string
import datetime
import shutil


def printline():
    print "\n======================================\n"
    return


def cleanUpPrompt(src1, src2):
    printline()

    choice = raw_input("Would you like the program to clear out the Source Files directory?\n" +\
              "Please enter the following commands to finish:\n\n" +\
              "Y/y = Yes (delete all files).\n" +\
              "N/n = No (do not make any changes to the Source Files directory).\n" +\
              "1 = Delete the 247 file.\n" +\
              "2 = Delete the Safari file.\n" +\
              "\nPlease enter your choice here: ")
    
    if choice == "Y" or choice == "y":
        os.remove(src1)
        os.remove(src2)
    elif choice == "N" or choice == "n":
        return
    elif choice.isdigit():
        if int(choice) == 1:
            os.remove(src1)
        elif int(choice) == 2:
            os.remove(src2)
    else:
        return cleanUpPrompt(src1, src2)
    
    return
        

def archiveFiles(src1, src2, srcFolder):
    now = datetime.datetime.now()
    
    archiveDir = os.path.join(srcFolder, "Archive/")
    bname247 = os.path.basename(src1)
    bnameSafari = os.path.basename(src2)
    
    shutil.copy(src1, archiveDir)
    shutil.copy(src2, archiveDir)

    os.rename(os.path.join(archiveDir, bname247), os.path.join(archiveDir, "247Archive" + now.strftime("%Y-%m-%d %H-%M") + ".xlsx"))
    os.rename(os.path.join(archiveDir, bnameSafari), os.path.join(archiveDir, "safariArchive" + now.strftime("%Y-%m-%d %H-%M") + ".xlsx"))

    return

def outputToFile(bmatch):
    now = datetime.datetime.now()
    wb = Workbook()

    ws = wb.get_active_sheet()

    ws.title = "Duplicate Records"

    ws.cell(row = 0, column = 0).value = "Record"
    ws.cell(row = 0, column = 1).value = "ISBN"
    ws.cell(row = 0, column = 2).value = "Title"
    ws.cell(row = 0, column = 3).value = "Author"
    ws.cell(row = 0, column = 4).value = "Edition"
    ws.cell(row = 0, column = 5).value = "Date"
    ws.cell(row = 0, column = 6).value = "URL"

    for i in range(1, len(bmatch)):
        temp_ISBNList = ""
        
        ws.cell(row = i, column = 0).value = bmatch[i].record

        
        # This next bit formats the ISBN numbers nicely, then
        # removes the extra ', ' from the end of the resulting string.
        
        for isbnRec in bmatch[i].isbn:
            temp_ISBNList += (isbnRec + ", ")

        temp_ISBNList = str(temp_ISBNList[:-2])
        
        ws.cell(row = i, column = 1).value = temp_ISBNList
        ws.cell(row = i, column = 2).value = bmatch[i].title
        ws.cell(row = i, column = 3).value = bmatch[i].author
        ws.cell(row = i, column = 4).value = bmatch[i].edition
        ws.cell(row = i, column = 5).value = bmatch[i].date
        ws.cell(row = i, column = 6).value = bmatch[i].url

    wb.save(os.path.join("Results", "result" + now.strftime("%Y-%m-%d %H:%M") + ".xlsx"))

    return    


def runBookMatching(blist, slist):
    bmatch = []
    
    for sbook in slist:
        for bbook in blist:
            if sbook == bbook:
                bmatch.append(sbook)
                
    return bmatch


def cleanUpISBN(src):
    x = []
    
    try:
        x = re.findall("\d+|[a-zA-Z]+", src.encode('ascii', 'ignore'))
        
    except AttributeError:
        try:
            return [str(int(src))]
        except TypeError:
            return ["none"]

    x = [item for item in x if item.isdigit()]

    return x


def findCol(colLoc, in_str):
    loc = ""
    
    for col in colLoc:
        if col[0].upper() == in_str:
            loc = string.lowercase.index(col[1].lower())
            
    return loc
            

def generateColLocations(src):
    colLocations = []
        
    for row in src.iter_rows('A1:L1'):
        for cell in row:
            if cell.internal_value != None:
                colLocations.append([cell.internal_value.encode('ascii', 'ignore'), cell.column])
                
    return colLocations


def buildBookDict(src):

    bookDict = []

    wb = load_workbook(filename = src, use_iterators = True)

    sheet = wb.worksheets[0]

    # generate an array that corresponds to the column.
    # [['RECORD #(BIBLIO)', 'A'], ['ISBN', 'B']]
    # values should be 'RECORD #(BIBLIO)', 'ISBN',
    # 'TITLE', 'AUTHOR', 'EDITION', 'DATE', 'URL'
    
    colLocations = generateColLocations(sheet)

    isbnCol = findCol(colLocations, 'ISBN')
    titleCol = findCol(colLocations, 'TITLE')
    authorCol = findCol(colLocations, 'AUTHOR')
    dateCol = findCol(colLocations, 'DATE')
    recordCol = findCol(colLocations, 'RECORD #(BIBLIO)')
    editionCol = findCol(colLocations, 'EDITION')
    urlCol = findCol(colLocations, 'URL')

    # so now we have the column IDs we need. Time to put the dictionary together
    
    for row in sheet.iter_rows():
        bookDict.append(LibraryBook(cleanUpISBN(row[isbnCol].internal_value), \
                                    row[recordCol].internal_value, \
                                    row[titleCol].internal_value, \
                                    row[authorCol].internal_value, \
                                    row[dateCol].internal_value,\
                                    row[editionCol].internal_value,\
                                    row[urlCol].internal_value))
         
    return bookDict
    
    

def selectCollectionFile(src):
    collFile = ""
    
    try:
        collFile = src[int(raw_input("\nPlease type your selection, then press return: ")) - 1]
        
    except IndexError:
        print "Invalid option. Please try again."
        collFile = selectCollectionFile(src)

    except ValueError:
        print "Invalid option. Please try again."
        collFile = selectCollectionFile(src)
        
    return collFile 


def printSourceDirList(src):
    for i in range(0, len(src)):
        print str(i + 1) + ". " + src[i]


def getSafari(src, chk, srcFld):
    printline()

    print "Thank you. Now, please select the Safari collection file\n" +\
          "from the list below. You will be prompted to try again if\n" +\
          "it matches your previous selection.\n"

    printSourceDirList(src)

    fname = selectCollectionFile(src)

    if fname == chk:
        print "Error: Safari collection name matches Books24x7 selection.\n" +\
              "Please try again."
        fname = getSafariColl(src, chk)

    return os.path.join(srcFld, fname)
    

def getBooks24x7(src, srcFld):

    print "The program found the following files in the resource folder.\n" + \
          "Please identify the file that represents the Books24x7 collection\n" + \
          "by entering the number below.\n"

    printSourceDirList(src)

    fname = selectCollectionFile(src)

    return os.path.join(srcFld, fname)
 

def main():
    sourceFolder = os.path.join(os.path.dirname(sys.argv[0]), "Source Files")

    sourceDirList = os.listdir(sourceFolder)

    books24x7 = getBooks24x7(sourceDirList, sourceFolder)

    safari = getSafari(sourceDirList, books24x7, sourceFolder)

    books24x7Dict = buildBookDict(books24x7)

    safariDict = buildBookDict(safari)

    bookMatch = runBookMatching(books24x7Dict, safariDict)

    printline()

    print "Identified " + str(len(bookMatch)) + " copies in the Safari database. " +\
          "Moving files and outputting to .xlsx now."

    outputToFile(bookMatch)

    archiveFiles(books24x7, safari, os.path.dirname(sys.argv[0]))

    cleanUpPrompt(books24x7, safari)
            

main()
